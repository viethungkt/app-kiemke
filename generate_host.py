import openpyxl
import json
import time
import sys

# Reconfigure stdout to use UTF-8 to prevent cp1252 encoding crashes on Windows
sys.stdout.reconfigure(encoding='utf-8')

# =========================================================
# CẤU HÌNH — Mapping từng Sheet sang Kho hàng tương ứng
# =========================================================
SHEET_CONFIG = [
    {
        "sheet":  "Nguyên Vật Liệu (NVL)",
        "kho":    "NL3",
        "prefix": "NVL",   # Tiền tố mã kiểm kê: NVL-1, NVL-2, ...
    },
    {
        "sheet":  "Thành Phẩm (TP)",
        "kho":    "XU",
        "prefix": "TP",    # TP-1, TP-2, ...
    },
    {
        "sheet":  "Phụ Tùng & Vật Tư (VT)",
        "kho":    "VPT",
        "prefix": "VT",    # VT-1, VT-2, ...
    },
    {
        "sheet":  "Bán Thành Phẩm & Phế Phẩm",
        "kho":    "XU",
        "prefix": "BTP",   # BTP-1, BTP-2, ...
    },
]

MASTER_PATH = "Danh_Muc_Vat_Tu_Kiem_Ke_Chot_23-06.xlsx"

# Cột trong mỗi sheet (dữ liệu bắt đầu từ row 5)
# A=1:STT  B=2:Mã cũ  C=3:Mã ERP Final  D=4:Tên chuẩn hóa
# E=5:DVT  F=6:Loại   G=7:Nhóm          H=8:Tồn sổ sách

def generate():
    start = time.time()

    # =========================================================
    # 1. Load danh sách kho từ Danh mục kho hàng vật tư.Xlsx
    # =========================================================
    khos = []
    try:
        print("Loading warehouse catalog from 'Danh mục kho hàng vật tư.Xlsx'...")
        wb_khos = openpyxl.load_workbook("Danh mục kho hàng vật tư.Xlsx", data_only=True)
        ws_khos = wb_khos.active
        for r in range(3, ws_khos.max_row + 1):
            makho  = ws_khos.cell(row=r, column=1).value
            tenkho = ws_khos.cell(row=r, column=2).value
            if makho and tenkho:
                khos.append({
                    "makho":  str(makho).strip(),
                    "tenkho": str(tenkho).strip()
                })
    except Exception as e:
        print(f"Error loading warehouses: {e}")
        khos = [
            {"makho": "XU",  "tenkho": "Kho Xưởng"},
            {"makho": "VPT", "tenkho": "Kho VP Trệt"},
            {"makho": "VP1", "tenkho": "Kho VP Lầu 1"},
            {"makho": "NL3", "tenkho": "Kho NL3"},
            {"makho": "KT",  "tenkho": "Kho Kỹ Thuật"},
        ]
    print(f"Loaded {len(khos)} warehouses.")

    # =========================================================
    # 2. Đọc tồn kho từng kho từ file cũ để map vị trí (vitri)
    #    File cũ: Danh Muc_TON KHO VAT TU THANG 23.06.xlsx
    # =========================================================
    OLD_STOCK_PATH = "Danh Muc_TON KHO VAT TU THANG 23.06.xlsx"
    old_sheet_to_makho = {
        "VP Lầu 1":  "VP1",
        "VP Trệt":   "VPT",
        "Kho Xuong": "XU",
        "Kho NL3":   "NL3",
        "Kho_KT":    "KT",
    }
    # name_lower -> {vitri, qty}
    old_vitri_map = {}

    def get_float(val):
        if val is None: return 0.0
        try: return float(val)
        except: return 0.0

    try:
        print(f"\nLoading old stock vitri from: {OLD_STOCK_PATH}")
        wb_old = openpyxl.load_workbook(OLD_STOCK_PATH, data_only=True)
        for s_name, makho in old_sheet_to_makho.items():
            if s_name not in wb_old.sheetnames:
                continue
            ws_old = wb_old[s_name]
            for r in range(6, ws_old.max_row + 1):
                stt_val = ws_old.cell(row=r, column=1).value
                if stt_val is None: continue
                name = ws_old.cell(row=r, column=2).value
                if not name: continue
                name_key = str(name).strip().lower()
                vitri = ws_old.cell(row=r, column=7).value or ""
                ke    = ws_old.cell(row=r, column=8).value or ""
                vitri = str(vitri).strip()
                ke    = str(ke).strip()
                if ke:
                    vitri = f"{vitri} - Kệ {ke}" if vitri else f"Kệ {ke}"
                if name_key not in old_vitri_map:
                    old_vitri_map[name_key] = vitri
        print(f"  Loaded vitri for {len(old_vitri_map)} unique item names.")
    except Exception as e:
        print(f"  Warning: Could not load old stock vitri: {e}")

    # =========================================================
    # 3. Import từng Sheet → Kho theo SHEET_CONFIG
    #    Mã kiểm kê = {PREFIX}-{STT}   (VD: NVL-1, TP-5, VT-12)
    # =========================================================
    print(f"\nLoading master catalog from: {MASTER_PATH}")
    wb_master = openpyxl.load_workbook(MASTER_PATH, data_only=True)

    items      = []   # list of item dicts
    book_stock = {}   # ma_erp -> ton_so_sach

    for cfg in SHEET_CONFIG:
        sheet_name = cfg["sheet"]
        makho      = cfg["kho"]
        prefix     = cfg["prefix"]

        if sheet_name not in wb_master.sheetnames:
            print(f"  WARNING: Sheet [{sheet_name}] not found, skipping.")
            continue

        ws = wb_master[sheet_name]
        sheet_count = 0

        for r in range(5, ws.max_row + 1):
            stt_val = ws.cell(row=r, column=1).value
            if stt_val is None:
                continue
            try:
                stt = int(float(str(stt_val)))
            except (ValueError, TypeError):
                continue

            ten_hang = ws.cell(row=r, column=4).value
            if not ten_hang:
                continue
            ten_hang = str(ten_hang).strip()

            ma_cu    = ws.cell(row=r, column=2).value
            ma_erp_f = ws.cell(row=r, column=3).value
            dvt      = ws.cell(row=r, column=5).value
            loai     = ws.cell(row=r, column=6).value
            nhom     = ws.cell(row=r, column=7).value
            ton_ss   = ws.cell(row=r, column=8).value  # Tồn sổ sách (thường None)

            ma_cu    = str(ma_cu).strip()    if ma_cu    else ""
            ma_erp_f = str(ma_erp_f).strip() if ma_erp_f else ma_cu
            dvt      = str(dvt).strip()      if dvt      else ""
            loai     = str(loai).strip()     if loai     else prefix
            nhom     = str(nhom).strip()     if nhom     else "Chưa phân nhóm"

            # ── Mã kiểm kê = PREFIX-STT (unique toàn hệ thống) ──
            ma_kiem_ke = f"{prefix}-{stt}"

            # ── Tồn sổ sách ──
            book_qty = get_float(ton_ss)

            # ── Vị trí: tra từ file cũ theo tên hàng ──
            vitri = old_vitri_map.get(ten_hang.lower(), "")

            items.append({
                "ma_erp":    ma_kiem_ke,   # Mã kiểm kê = số thứ tự có tiền tố
                "ma_qr":     ma_kiem_ke,   # QR barcode cũng dùng mã này
                "ten_hang":  ten_hang,
                "ten_cu":    ma_cu,        # Mã cũ / barcode cũ lưu ở đây
                "dvt":       dvt,
                "nhom":      nhom,
                "vitri":     vitri,
                "loai":      loai,
                "ma_cu":     ma_cu,        # Mã cũ để tra cứu / tìm kiếm
                "ma_erp_full": ma_erp_f,   # Mã ERP Full để tham chiếu
                "kho":       makho,
            })
            book_stock[ma_kiem_ke] = book_qty
            sheet_count += 1

        print(f"  Sheet [{sheet_name}] → Kho [{makho}]: {sheet_count} items (mã: {prefix}-1 → {prefix}-{sheet_count})")

    # =========================================================
    # 4. Thống kê
    # =========================================================
    print(f"\nTotal items: {len(items)}")
    print(f"Total book_stock entries: {len(book_stock)}")

    non_zero_stock = sum(1 for v in book_stock.values() if v > 0)
    print(f"Book stock non-zero: {non_zero_stock}")

    kho_dist = {}
    for item in items:
        kho_dist[item["kho"]] = kho_dist.get(item["kho"], 0) + 1
    print("Distribution by warehouse:")
    for k, v in sorted(kho_dist.items()):
        print(f"  {k}: {v} items")

    # Sample
    print("\nSample (first 3 items per sheet):")
    seen = set()
    shown = 0
    for item in items:
        prefix_key = item["ma_erp"].split("-")[0]
        if prefix_key not in seen and shown < 12:
            seen.add(prefix_key)
            print(f"  {item['ma_erp']} | {item['kho']} | {item['ten_hang'][:50]} | ma_cu={item['ma_cu']}")
            shown += 1

    # =========================================================
    # 5. Ghi ra file JS
    # =========================================================
    with open("data_kho.js", "w", encoding="utf-8") as f:
        f.write("/* DATA DANH MỤC KHO HÀNG HUỆ LINH */\n")
        f.write(f"const KHO_DEFAULT = {json.dumps(khos, ensure_ascii=False, indent=2)};\n")
        f.write('if (typeof window !== "undefined") { window.KHO_DEFAULT = KHO_DEFAULT; }\n')
        f.write('if (typeof module !== "undefined" && module.exports) { module.exports = { KHO_DEFAULT }; }\n')

    with open("data_vattu.js", "w", encoding="utf-8") as f:
        f.write("/* DATA DANH MỤC VẬT TƯ — Nguồn: Danh_Muc_Vat_Tu_Kiem_Ke_Chot_23-06.xlsx */\n")
        f.write("/* Mã kiểm kê = PREFIX-STT (NVL-1, TP-1, VT-1, BTP-1) */\n")
        f.write(f"const VAT_TU_DEFAULT = {json.dumps(items, ensure_ascii=False, indent=2)};\n\n")
        f.write(f"const BOOK_STOCK_DEFAULT = {json.dumps(book_stock, ensure_ascii=False, indent=2)};\n\n")
        f.write('if (typeof window !== "undefined") {\n')
        f.write('  window.VAT_TU_DEFAULT = VAT_TU_DEFAULT;\n')
        f.write('  window.BOOK_STOCK_DEFAULT = BOOK_STOCK_DEFAULT;\n')
        f.write('}\n')
        f.write('if (typeof module !== "undefined" && module.exports) { module.exports = { VAT_TU_DEFAULT, BOOK_STOCK_DEFAULT }; }\n')

    elapsed = time.time() - start
    print(f"\n✓ Generated data_kho.js and data_vattu.js in {elapsed:.2f}s")


if __name__ == "__main__":
    generate()
