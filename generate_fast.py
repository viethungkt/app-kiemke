import openpyxl
import json
import time
import sys

# Reconfigure stdout to use UTF-8 to prevent cp1252 encoding crashes on Windows
sys.stdout.reconfigure(encoding='utf-8')

def generate():
    start = time.time()
    
    # Load warehouse list from Danh mục kho hàng vật tư.Xlsx
    khos = []
    try:
        print("Loading warehouse catalog from 'Danh mục kho hàng vật tư.Xlsx'...")
        wb_khos = openpyxl.load_workbook("Danh mục kho hàng vật tư.Xlsx", data_only=True)
        ws_khos = wb_khos.active
        for r in range(3, ws_khos.max_row + 1):
            makho = ws_khos.cell(row=r, column=1).value
            tenkho = ws_khos.cell(row=r, column=2).value
            if makho and tenkho:
                khos.append({
                    "makho": str(makho).strip(),
                    "tenkho": str(tenkho).strip()
                })
    except Exception as e:
        print(f"Error loading warehouses: {e}")
        khos = [
            {"makho": "XU", "tenkho": "Kho Xuong"},
            {"makho": "VPT", "tenkho": "KHO VP Trệt"},
            {"makho": "VP1", "tenkho": "KHO VP Lầu 1"},
            {"makho": "NL3", "tenkho": "KHO Kho NL3"},
            {"makho": "KT", "tenkho": "KHO Kỹ Thuật"}
        ]
    print(f"Loaded {len(khos)} warehouses from catalog.")

    # Load master catalog for enrichment
    master_path = "Danh_Muc_Vat_Tu_Kiem_Ke_Chot_23-06.xlsx"
    master_map = {}
    try:
        print("Loading master catalog for enrichment...")
        wb_master = openpyxl.load_workbook(master_path, data_only=True)
        ws_master = wb_master["Tất cả danh mục"]
        for r in range(5, ws_master.max_row + 1):
            name = ws_master.cell(row=r, column=4).value
            if name:
                name_norm = str(name).strip().lower()
                master_map[name_norm] = {
                    "nhom": ws_master.cell(row=r, column=7).value,
                    "loai": ws_master.cell(row=r, column=6).value,
                    "ma_cu": ws_master.cell(row=r, column=2).value
                }
        print(f"Loaded master map: {len(master_map)} items.")
    except Exception as e:
        print(f"Warning: Could not load master catalog for enrichment: {e}")

    sheet_to_makho = {
        "VP Lầu 1": "VP1",
        "VP Trệt": "VPT",
        "Kho Xuong": "XU",
        "Kho NL3": "NL3",
        "Kho_KT": "KT"
    }

    # Load stock from Danh Muc_TON KHO VAT TU THANG 23.06.xlsx
    items_dict = {}
    book_stock = {}

    def get_stt_int(val):
        if val is None:
            return None
        try:
            return int(val)
        except (ValueError, TypeError):
            try:
                f = float(val)
                return int(f)
            except (ValueError, TypeError):
                return None

    def get_qty(val):
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    stock_path = "Danh Muc_TON KHO VAT TU THANG 23.06.xlsx"
    try:
        print("Loading stock data from 'Danh Muc_TON KHO VAT TU THANG 23.06.xlsx'...")
        wb_stock = openpyxl.load_workbook(stock_path, data_only=True)
        
        # Process sheets in order from general to specific
        sheets_order = ["VP Lầu 1", "VP Trệt", "Kho Xuong", "Kho NL3", "Kho_KT"]
        for s_name in sheets_order:
            if s_name not in wb_stock.sheetnames:
                continue
            ws = wb_stock[s_name]
            makho = sheet_to_makho[s_name]
            sheet_count = 0
            
            for r in range(6, ws.max_row + 1):
                stt_val = ws.cell(row=r, column=1).value
                stt = get_stt_int(stt_val)
                if stt is None:
                    continue
                
                name = ws.cell(row=r, column=2).value
                if not name:
                    continue
                name = str(name).strip()
                
                dvt = ws.cell(row=r, column=6).value
                dvt = str(dvt).strip() if dvt else ""
                
                vitri = ws.cell(row=r, column=7).value
                vitri = str(vitri).strip() if vitri else ""
                
                ke = ws.cell(row=r, column=8).value
                ke = str(ke).strip() if ke else ""
                
                if ke:
                    vitri = f"{vitri} - Kệ {ke}" if vitri else f"Kệ {ke}"
                    
                qty_val = get_qty(ws.cell(row=r, column=9).value)
                
                name_norm = name.lower()
                if name_norm in master_map:
                    nhom = master_map[name_norm]["nhom"]
                    loai = master_map[name_norm]["loai"]
                    ma_cu = master_map[name_norm]["ma_cu"]
                else:
                    nhom = f"Vật tư {s_name}"
                    loai = "VT"
                    ma_cu = ""
                    
                if stt not in items_dict:
                    items_dict[stt] = {
                        "ma_erp": str(stt),
                        "ma_qr": str(stt),
                        "ten_hang": name,
                        "ten_cu": name,
                        "dvt": dvt,
                        "nhom": nhom if nhom else "Vật tư",
                        "vitri": vitri,
                        "loai": loai if loai else "VT",
                        "ma_cu": ma_cu if ma_cu else "",
                        "kho": makho
                    }
                    book_stock[str(stt)] = qty_val
                    sheet_count += 1
                else:
                    # Update existing duplicate STT item with specific warehouse data
                    items_dict[stt]["kho"] = makho
                    if vitri:
                        items_dict[stt]["vitri"] = vitri
                    book_stock[str(stt)] = qty_val
                    
            print(f"Processed sheet {s_name}: Added {sheet_count} new unique items.")
    except Exception as e:
        print(f"Error loading stock info: {e}")

    items = list(items_dict.values())
    print(f"Total unique items: {len(items)}")
    print(f"Total book stock entries: {len(book_stock)}")
    
    # Write directly to workspace
    with open("data_kho.js", "w", encoding="utf-8") as f:
        f.write("/* DATA DANH MỤC KHO HÀNG HUÊ LINH */\n")
        f.write(f"const KHO_DEFAULT = {json.dumps(khos, ensure_ascii=False, indent=2)};\n")
        f.write("if (typeof window !== \"undefined\") { window.KHO_DEFAULT = KHO_DEFAULT; }\n")
        f.write("if (typeof module !== \"undefined\" && module.exports) { module.exports = { KHO_DEFAULT }; }\n")
        
    with open("data_vattu.js", "w", encoding="utf-8") as f:
        f.write("/* DATA DANH MỤC VẬT TƯ HÀNG HÓA HUÊ LINH */\n")
        f.write(f"const VAT_TU_DEFAULT = {json.dumps(items, ensure_ascii=False, indent=2)};\n\n")
        f.write(f"const BOOK_STOCK_DEFAULT = {json.dumps(book_stock, ensure_ascii=False, indent=2)};\n\n")
        f.write("if (typeof window !== \"undefined\") {\n")
        f.write("  window.VAT_TU_DEFAULT = VAT_TU_DEFAULT;\n")
        f.write("  window.BOOK_STOCK_DEFAULT = BOOK_STOCK_DEFAULT;\n")
        f.write("}\n")
        f.write("if (typeof module !== \"undefined\" && module.exports) { module.exports = { VAT_TU_DEFAULT, BOOK_STOCK_DEFAULT }; }\n")

    print(f"Successfully generated data files in {time.time() - start:.3f}s")

if __name__ == "__main__":
    generate()
